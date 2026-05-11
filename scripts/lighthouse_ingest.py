"""lighthouse_ingest.py — XLSX → long-format CSV preprocessing for Hyatt Regency San Francisco Embarcadero.

Reads Lighthouse Rate Insight `.xlsx` exports from the deal's
`Lighthouse/Drops/` directory and writes a single long-format
`lighthouse_rates.csv` consumed by `build_dashboard.py`.

Run before each `redeploy.py` invocation:

    py scripts/lighthouse_ingest.py --all

The CSV is committable as a frozen snapshot — it does NOT auto-rebuild from
the XLSX. Re-run this script when a new export drops.

Per-deal contract (Resolution 12, 14, 15)
-----------------------------------------
* `lighthouse_subject_hotel_id` = 182230
* `lighthouse_host_account_hotel_id` = 284844
* Channel = `booking` — written into the `source` column.
* Sentinel taxonomy: lifted from `scraper_lib/sentinel.py` (Resolution 15).
  The Lighthouse `Rates` sheet interleaves numeric rates with string
  sentinels (`Sold out`, `No flex`, `LOS2`, `--`, etc.); `normalize_cell`
  maps each cell to (rate_usd, availability_status, los_restriction).

Output schema (12 columns)
--------------------------
    as_of_date, arrival_date, dow, source, room_tier, property,
    rate_usd, availability_status, los_restriction,
    market_demand_frac, market_otb_frac, length_of_stay_nights

`length_of_stay_nights` is captured from the source filename — see
`FILENAME_LOS_RE`. Single-LOS deals will see a constant value; multi-LOS
deals (SFOEM-pattern: 3 files for LOS=1, 3, 7) will see the discriminator
populated so downstream analytics can filter.

Drop invariants
---------------
The host-account column and `My OTB` are dropped at parser layer — they
contaminate cross-property analysis if left in. Drop list is pre-rendered
from `lighthouse_drop_columns` and always includes `My OTB`.
"""
from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

# Make scraper_lib importable when running from the repo root.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scraper_lib"))
from sentinel import (  # noqa: E402
    LighthouseSlugMapDriftError,
    normalize_cell,
    to_float_or_none,
    validate_lighthouse_slug_map_coverage,
)

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger("lighthouse_ingest")

# ---------------------------------------------------------------------------
# Per-deal configuration (rendered from skill scaffold parameters).
# ---------------------------------------------------------------------------

# Property header (verbatim from the Lighthouse XLSX) → canonical slug.
# Headers not in this map are skipped with a warning. Subject's slug must
# match `subject_slug` in analytics_lighthouse.py.
LIGHTHOUSE_PROPERTY_SLUG: dict[str, str] = {
    "Hyatt Regency San Francisco": "hr_embarcadero",
    "Hyatt Regency San Francisco Downtown SOMA": "hr_soma",
    "The Clancy, Autograph Collection": "clancy",
    "Hilton San Francisco Union Square": "hilton_us",
    "Grand Hyatt San Francisco Union Square": "grand_hyatt",
    "Palace Hotel, a Luxury Collection Hotel, San Francisco": "palace",
    "San Francisco Marriott Marquis Union Square": "marquis",
    "InterContinental San Francisco by IHG": "ic_sf",
    "The Westin St. Francis San Francisco on Union Square": "st_francis",
}

# Headers to drop unconditionally — host account (Resolution 12) + My OTB.
LIGHTHOUSE_DROP_COLUMNS: list[str] = [
    "Park Central Hotel New York",
]

# Channel the panel reports on (Resolution 13). Written into `source`.
LIGHTHOUSE_CHANNEL: str = "booking"

# Sheet inside the XLSX (Resolution 14). Default `Rates`.
SHEET_NAME: str = "Rates"

# Filename glob (Resolution 14). LOS captured from filename via FILENAME_LOS_RE.
FILENAME_GLOB: str = "park-central-hotel-new-york_bookingdotcom_bar_los*_2guests_sec.xlsx"

# Regex with a named (?P<los>\d+) group OR named (?P<as_of>YYYY-MM-DD).
# LOS defaults to 1 when no group matches; as_of defaults to file mtime date.
FILENAME_LOS_RE = re.compile(r"_los(?P<los>\d+)_", re.IGNORECASE)
FILENAME_ASOF_RE = re.compile(r"(?P<as_of>\d{4}-\d{2}-\d{2})")

# Output path — auto-derived (Resolution 10) but overridable via --out.
# Lighthouse/ is a sibling of the repo, not a child of a property-named folder.
DEFAULT_OUT_PATH = (
    Path(__file__).resolve().parent.parent.parent
    / "Lighthouse"
    / "lighthouse_rates.csv"
)

# ---------------------------------------------------------------------------
# Excel layout — Lighthouse Rate Insight standard.
# ---------------------------------------------------------------------------

EXCEL_HEADER_ROW = 4          # pandas header= argument (0-indexed); openpyxl row 5
EXCEL_DATA_FIRST_ROW = 5      # pandas 0-indexed; openpyxl row 6
DATE_ORIGIN = "1899-12-30"

# Lighthouse Rate Insight stamps the data-as-of timestamp at row 3 of the
# Rates sheet: label "Updated" at column F, value "DD/MM HH:MM" at column G.
# Day-first per Lighthouse's reporting locale.
EXCEL_UPDATED_ROW = 3
EXCEL_UPDATED_LABEL_COL = 6
EXCEL_UPDATED_VALUE_COL = 7
WORKBOOK_UPDATED_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})\s+\d{1,2}:\d{2}\s*$")

DAY_HEADER = "Day"
DATE_HEADER = "Date"
MY_OTB_HEADER = "My OTB"
MARKET_OTB_HEADER = "Market OTB"
MARKET_DEMAND_HEADER = "Market demand"

# Schema-assertion bounds.
EXPECTED_DATE_HORIZON_MIN = 360
EXPECTED_DATE_HORIZON_MAX = 370

OUTPUT_COLUMNS = [
    "as_of_date", "arrival_date", "dow", "source", "room_tier", "property",
    "rate_usd", "availability_status", "los_restriction",
    "market_demand_frac", "market_otb_frac", "length_of_stay_nights",
]


@dataclass
class FilenameMeta:
    los: int
    as_of_date: pd.Timestamp


def parse_filename(path: Path) -> FilenameMeta:
    """Extract LOS + as-of date from filename via configured regexes."""
    los_m = FILENAME_LOS_RE.search(path.name)
    los = int(los_m.group("los")) if los_m and "los" in los_m.groupdict() else 1

    asof_m = FILENAME_ASOF_RE.search(path.name)
    if asof_m and "as_of" in asof_m.groupdict():
        as_of = pd.to_datetime(asof_m.group("as_of"), errors="coerce")
    else:
        as_of = pd.NaT
    if pd.isna(as_of):
        as_of = pd.to_datetime(path.stat().st_mtime, unit="s").normalize()
    return FilenameMeta(los=los, as_of_date=as_of)


def _read_workbook_updated_asof(path: Path) -> Optional[pd.Timestamp]:
    """Read the 'Updated' cell from the Rates sheet (F3 label, G3 value).

    Lighthouse exports without a date in the filename carry the data-as-of
    timestamp here; trusting it is more truthful than file mtime, which can
    drift by days between export and drop-folder placement. Returns the
    date portion (normalized) or None if absent/unparseable; caller falls
    back to mtime.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            if SHEET_NAME not in wb.sheetnames:
                return None
            ws = wb[SHEET_NAME]
            label = ws.cell(row=EXCEL_UPDATED_ROW, column=EXCEL_UPDATED_LABEL_COL).value
            value = ws.cell(row=EXCEL_UPDATED_ROW, column=EXCEL_UPDATED_VALUE_COL).value
        finally:
            wb.close()
    except (FileNotFoundError, OSError, ValueError) as exc:
        log.warning("%s: failed reading Updated cell: %s", path.name, exc)
        return None

    if not isinstance(label, str) or label.strip().lower() != "updated":
        return None
    if not isinstance(value, str):
        return None
    m = WORKBOOK_UPDATED_RE.match(value)
    if not m:
        log.warning("%s: 'Updated' cell unparseable: %r", path.name, value)
        return None
    day, month = int(m.group(1)), int(m.group(2))
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None

    # No year in the cell — anchor to the file's mtime year, then snap back
    # one year if the candidate is implausibly far in the future (handles
    # exports straddling Jan 1).
    mtime = pd.to_datetime(path.stat().st_mtime, unit="s")
    try:
        candidate = pd.Timestamp(year=mtime.year, month=month, day=day)
    except ValueError:
        return None
    if candidate > mtime + pd.Timedelta(days=90):
        try:
            candidate = pd.Timestamp(year=mtime.year - 1, month=month, day=day)
        except ValueError:
            return None
    return candidate.normalize()


def _excel_serial_to_date(serial) -> Optional[pd.Timestamp]:
    if serial is None:
        return None
    if isinstance(serial, pd.Timestamp):
        return serial
    if isinstance(serial, (int, float)):
        if pd.isna(serial) or serial == 0:
            return None
        return pd.to_datetime(serial, unit="D", origin=DATE_ORIGIN)
    try:
        return pd.to_datetime(serial)
    except (ValueError, TypeError):
        return None


def _finalize_output_frame(df: pd.DataFrame) -> pd.DataFrame:
    df = df.reindex(columns=OUTPUT_COLUMNS)
    for col in ("rate_usd", "market_demand_frac", "market_otb_frac"):
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Float64")
    for col in ("los_restriction", "length_of_stay_nights"):
        df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
    return df


# ---------------------------------------------------------------------------
# Per-file parser.
# ---------------------------------------------------------------------------

def parse_drop_file(path: Path) -> pd.DataFrame:
    """Parse one Lighthouse XLSX → long-format DataFrame.

    Drops `LIGHTHOUSE_DROP_COLUMNS` headers at parser layer (single invariant).
    Truncates trailing empty/zero rows (Lighthouse pads the sheet beyond the
    365-day forward window).
    """
    meta = parse_filename(path)
    # Prefer the workbook 'Updated' cell when the filename had no date —
    # mtime can lag the actual export by days (LOS7 SFOEM dry-run: cell
    # 2026-05-06, mtime 2026-05-08).
    if not FILENAME_ASOF_RE.search(path.name):
        workbook_asof = _read_workbook_updated_asof(path)
        if workbook_asof is not None:
            meta = FilenameMeta(los=meta.los, as_of_date=workbook_asof)

    df = pd.read_excel(path, sheet_name=SHEET_NAME, header=EXCEL_HEADER_ROW, engine="openpyxl")

    headers = list(df.columns)
    for required in (DATE_HEADER, MARKET_DEMAND_HEADER, MARKET_OTB_HEADER):
        if required not in headers:
            raise ValueError(
                f"{path.name}: missing '{required}' column. Headers: {headers}"
            )

    metadata_cols = {DAY_HEADER, DATE_HEADER, MY_OTB_HEADER, MARKET_OTB_HEADER, MARKET_DEMAND_HEADER}

    # Finding 26 fail-loud (Resolution 20). Every XLSX column must
    # resolve to either the slug map, the drop list, or known metadata —
    # silent skips would drop properties from the long-format CSV
    # without a trace. SFOEM dry-run found 5 of 9 drafted slug-map keys
    # would have failed silent dict lookups against the actual XLSX
    # headers due to verbatim-string drift (uppercase, suffix, brand-tag
    # variants). Raise at parse time so the analyst reconciles the slug
    # map against the export before the dashboard is built.
    validate_lighthouse_slug_map_coverage(
        headers,
        slug_map=LIGHTHOUSE_PROPERTY_SLUG,
        drop_columns=set(LIGHTHOUSE_DROP_COLUMNS) | {MY_OTB_HEADER},
        metadata_cols=metadata_cols,
    )

    drop_seen: set[str] = set()
    property_columns: List[str] = []
    for h in headers:
        if not isinstance(h, str):
            continue
        if h in metadata_cols:
            continue
        if h.startswith("Unnamed:"):
            continue
        if h in LIGHTHOUSE_DROP_COLUMNS or h == MY_OTB_HEADER:
            drop_seen.add(h)
            continue
        property_columns.append(h)

    expected_drops = set(LIGHTHOUSE_DROP_COLUMNS)
    missing_drops = expected_drops - drop_seen
    if missing_drops:
        log.warning(
            "%s: configured drop column(s) not found in headers: %s. "
            "Drop invariant cannot be confirmed. Headers=%s",
            path.name, sorted(missing_drops), headers,
        )

    unknown = [h for h in property_columns if h not in LIGHTHOUSE_PROPERTY_SLUG]
    if unknown:
        log.warning("%s: unknown property headers (skipped): %s", path.name, unknown)
        property_columns = [h for h in property_columns if h in LIGHTHOUSE_PROPERTY_SLUG]

    records: List[dict] = []
    unknown_sentinels: Dict[str, int] = {}
    truncated_count = 0

    for _, row in df.iterrows():
        arrival = _excel_serial_to_date(row[DATE_HEADER])
        if arrival is None:
            truncated_count += 1
            continue

        dow_raw = row.get(DAY_HEADER)
        dow = dow_raw.strip() if isinstance(dow_raw, str) else arrival.strftime("%a")

        market_demand_frac = to_float_or_none(row.get(MARKET_DEMAND_HEADER))
        market_otb_frac = to_float_or_none(row.get(MARKET_OTB_HEADER))

        for header in property_columns:
            slug = LIGHTHOUSE_PROPERTY_SLUG[header]
            rate_usd, status, los = normalize_cell(row[header], unknown_sentinels)
            records.append({
                "as_of_date": meta.as_of_date.date(),
                "arrival_date": arrival.date(),
                "dow": dow,
                "source": LIGHTHOUSE_CHANNEL,
                "room_tier": "any",
                "property": slug,
                "rate_usd": rate_usd,
                "availability_status": status,
                "los_restriction": los,
                "market_demand_frac": market_demand_frac,
                "market_otb_frac": market_otb_frac,
                "length_of_stay_nights": meta.los,
            })

    out = _finalize_output_frame(pd.DataFrame.from_records(records, columns=OUTPUT_COLUMNS))

    not_loaded_by_property = (
        out[out["availability_status"] == "not_loaded"]["property"]
        .value_counts().to_dict()
    )
    sold_out_by_property = (
        out[out["availability_status"] == "sold_out"]["property"]
        .value_counts().to_dict()
    )
    out.attrs["not_loaded_count_by_property"] = not_loaded_by_property
    out.attrs["sold_out_count_by_property"] = sold_out_by_property
    out.attrs["truncated_row_count"] = truncated_count
    out.attrs["unknown_sentinels"] = dict(unknown_sentinels)

    if truncated_count:
        log.info("%s: truncated %d trailing/empty rows.", path.name, truncated_count)
    if unknown_sentinels:
        log.warning(
            "%s: unrecognized cell strings (counts): %s. "
            "Add to scraper_lib/sentinel.py if these are stable Lighthouse strings.",
            path.name, unknown_sentinels,
        )

    return out


def parse_all_drops(drops_dir: Path) -> pd.DataFrame:
    files = sorted(p for p in drops_dir.glob(FILENAME_GLOB) if p.is_file())
    if not files:
        raise FileNotFoundError(
            f"No files matching '{FILENAME_GLOB}' in {drops_dir}. "
            "Drop the Lighthouse export(s) into this directory and re-run."
        )

    dfs: List[pd.DataFrame] = []
    for f in files:
        log.info("Parsing %s", f.name)
        dfs.append(parse_drop_file(f))

    main_df = (
        pd.concat([d for d in dfs if not d.empty], ignore_index=True)
        if any(not d.empty for d in dfs)
        else _finalize_output_frame(pd.DataFrame(columns=OUTPUT_COLUMNS))
    )
    return main_df


def write_long_csv(df: pd.DataFrame, out: Path) -> None:
    out.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out, index=False)
    log.info("Wrote %s rows × %s cols → %s", df.shape[0], df.shape[1], out)


# ---------------------------------------------------------------------------
# Anchor-validate — sample N available rows; confirm rate matches the source.
# ---------------------------------------------------------------------------

def anchor_validate(
    df: pd.DataFrame, source_xlsx: Path, n: int = 20, seed: int = 42,
) -> int:
    """Pick n random `available` rows; walk source XLSX with openpyxl;
    confirm rate matches. Returns mismatch count."""
    available = df[df["availability_status"] == "available"]
    if len(available) < n:
        log.warning("Only %d available rows; sampling all.", len(available))
        n = len(available)
    if n == 0:
        return 0
    sample = available.sample(n=n, random_state=seed)

    wb = openpyxl.load_workbook(source_xlsx, data_only=True, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"{source_xlsx}: no '{SHEET_NAME}' sheet")
    ws = wb[SHEET_NAME]

    header_row_excel = EXCEL_HEADER_ROW + 1
    header_cells = [c.value for c in ws[header_row_excel]]
    header_to_col = {
        (h.strip() if isinstance(h, str) else None): i + 1
        for i, h in enumerate(header_cells)
    }
    date_col = header_to_col.get(DATE_HEADER)
    if date_col is None:
        raise ValueError(f"{source_xlsx}: no Date column")
    property_to_col: Dict[str, int] = {}
    for h, idx in header_to_col.items():
        if isinstance(h, str) and h in LIGHTHOUSE_PROPERTY_SLUG:
            property_to_col[LIGHTHOUSE_PROPERTY_SLUG[h]] = idx

    date_to_row: Dict[object, int] = {}
    for excel_row in range(EXCEL_DATA_FIRST_ROW + 1, ws.max_row + 1):
        cell = ws.cell(row=excel_row, column=date_col).value
        d = _excel_serial_to_date(cell)
        if d is not None:
            date_to_row[d.date()] = excel_row

    mismatches = 0
    for _, parsed_row in sample.iterrows():
        prop = parsed_row["property"]
        arrival = parsed_row["arrival_date"]
        expected = parsed_row["rate_usd"]
        if prop not in property_to_col:
            mismatches += 1
            continue
        excel_row = date_to_row.get(arrival)
        if excel_row is None:
            mismatches += 1
            continue
        cell_val = ws.cell(row=excel_row, column=property_to_col[prop]).value
        if not isinstance(cell_val, (int, float)) or cell_val is None:
            mismatches += 1
            continue
        if abs(float(cell_val) - float(expected)) > 1e-6:
            log.warning(
                "anchor_validate MISMATCH: prop=%s arrival=%s parsed=%s xlsx=%s",
                prop, arrival, expected, cell_val,
            )
            mismatches += 1

    wb.close()
    log.info("Anchor-validator: %d / %d rows match", n - mismatches, n)
    return mismatches


# ---------------------------------------------------------------------------
# Diagnostics + CLI.
# ---------------------------------------------------------------------------

def _print_diagnostics(df: pd.DataFrame, label: str) -> None:
    print(f"\n=== Diagnostics for {label} ===")
    print(f"shape: {df.shape}")
    print(f"length_of_stay_nights: {sorted(df['length_of_stay_nights'].dropna().unique().tolist())}")
    print("\nvalue_counts(availability_status):")
    print(df["availability_status"].value_counts().to_string())
    unknown = df[df["availability_status"].astype(str).str.startswith("unknown_sentinel:")]
    if not unknown.empty:
        print(f"\nUNKNOWN sentinel rows: {len(unknown)}  [WARN]")
        print(unknown["availability_status"].value_counts().to_string())
    else:
        print("\nUnknown sentinels: 0  [OK]")


def main() -> int:
    parser = argparse.ArgumentParser(description="Lighthouse XLSX → long-format CSV.")
    parser.add_argument(
        "--drops-dir", type=Path,
        default=Path(__file__).resolve().parent.parent.parent
            / "Lighthouse" / "Drops",
    )
    parser.add_argument("--single", type=Path, default=None,
                        help="Parse a single XLSX; print diagnostics; do not write CSV.")
    parser.add_argument("--all", action="store_true",
                        help="Parse all matching files and write the canonical CSV.")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT_PATH)
    parser.add_argument("--anchor-n", type=int, default=20)
    args = parser.parse_args()

    if args.single:
        df = parse_drop_file(args.single)
        _print_diagnostics(df, args.single.name)
        mismatches = anchor_validate(df, args.single, n=args.anchor_n)
        if mismatches:
            print(f"FAIL: {mismatches} anchor-validate mismatches")
            return 1
        print("OK")
        return 0

    if args.all:
        df = parse_all_drops(args.drops_dir)
        write_long_csv(df, args.out)
        _print_diagnostics(df, str(args.out))
        return 0

    parser.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
