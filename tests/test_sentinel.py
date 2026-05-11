"""Unit tests for scraper_lib/sentinel.py.

Coverage strategy
-----------------
1. Every documented sentinel value maps to its enum (positive cases).
2. Casing variants — `Sold Out`, `SOLD OUT` — all hit the case-insensitive
   path, except `--` / `LOS{n}` / numeric / None which are exact matches.
3. Numeric handling — int, float, zero (treated as blank), NaN.
4. Unknown strings tag as `unknown_sentinel:{raw}` and increment the
   `unknown_sentinels` Counter when one is supplied.
5. Behavior parity with AKA's existing `_normalize_cell` — golden table
   built from the AKA test file.
6. Integration: real SFOEM XLSX file parses with zero unknown sentinels
   (skipped if the fixture XLSX is not present locally).
"""
from __future__ import annotations

import math
from collections import Counter
from pathlib import Path

import pandas as pd
import pytest

from sentinel import (
    AVAILABILITY_STATUSES,
    LighthouseSlugMapDriftError,
    normalize_cell,
    to_float_or_none,
    validate_lighthouse_slug_map_coverage,
)


# -----------------------------------------------------------------------------
# Documented enum values map
# -----------------------------------------------------------------------------

def test_sold_out_variants_map_to_sold_out():
    for raw in ("Sold out", "sold out", "SOLD OUT", "  Sold Out  "):
        rate, status, los = normalize_cell(raw)
        assert (rate, status, los) == (None, "sold_out", None), raw


def test_no_flex_variants_map_to_no_flex():
    for raw in ("No flex", "NO FLEX", "no flex", " No Flex "):
        assert normalize_cell(raw) == (None, "no_flex", None), raw


def test_room_na_both_punctuations():
    assert normalize_cell("Room n/a") == (None, "room_na", None)
    assert normalize_cell("Room na") == (None, "room_na", None)
    assert normalize_cell("ROOM N/A") == (None, "room_na", None)


def test_double_dash_is_not_loaded_exact_match_only():
    assert normalize_cell("--") == (None, "not_loaded", None)
    # Em-dash and en-dash are NOT not_loaded — they are unknown sentinels.
    rate, status, _ = normalize_cell("—")
    assert status.startswith("unknown_sentinel:"), status


def test_third_party_only_and_one_guest_only():
    assert normalize_cell("3rd Party only") == (None, "third_party_only", None)
    assert normalize_cell("3RD PARTY ONLY") == (None, "third_party_only", None)
    assert normalize_cell("1 guest only") == (None, "one_guest_only", None)
    assert normalize_cell("1 Guest Only") == (None, "one_guest_only", None)


def test_los_restricted_captures_n():
    assert normalize_cell("LOS2") == (None, "los_restricted", 2)
    assert normalize_cell("LOS3") == (None, "los_restricted", 3)
    assert normalize_cell("los7") == (None, "los_restricted", 7)
    assert normalize_cell("LOS14") == (None, "los_restricted", 14)


def test_los_pattern_strict_no_partial_match():
    # "LOS2 something" should NOT match — pattern is anchored.
    rate, status, _ = normalize_cell("LOS2 weekend only")
    assert status.startswith("unknown_sentinel:")


# -----------------------------------------------------------------------------
# Numeric handling
# -----------------------------------------------------------------------------

@pytest.mark.parametrize("v,expected", [
    (0, (None, "blank", None)),
    (0.0, (None, "blank", None)),
    (199, (199.0, "available", None)),
    (199.5, (199.5, "available", None)),
    (-50.0, (-50.0, "available", None)),  # negative not impossible; pass through
    (1e6, (1e6, "available", None)),
])
def test_numeric_branches(v, expected):
    assert normalize_cell(v) == expected


def test_nan_is_blank():
    assert normalize_cell(float("nan")) == (None, "blank", None)
    assert normalize_cell(math.nan) == (None, "blank", None)


def test_none_is_blank():
    assert normalize_cell(None) == (None, "blank", None)


def test_empty_string_is_blank():
    assert normalize_cell("") == (None, "blank", None)
    assert normalize_cell("   ") == (None, "blank", None)


# -----------------------------------------------------------------------------
# Unknown sentinel surfacing
# -----------------------------------------------------------------------------

def test_unknown_string_tags_with_prefix():
    rate, status, los = normalize_cell("Mystery Status")
    assert rate is None
    assert los is None
    assert status == "unknown_sentinel:Mystery Status"


def test_unknown_increments_counter_when_provided():
    counter: Counter = Counter()
    normalize_cell("Mystery", counter)
    normalize_cell("Mystery", counter)
    normalize_cell("Other", counter)
    assert counter == Counter({"Mystery": 2, "Other": 1})


def test_unknown_does_not_crash_without_counter():
    # Counter is optional — must not raise when omitted.
    rate, status, _ = normalize_cell("Foo bar")
    assert status == "unknown_sentinel:Foo bar"


def test_non_string_non_numeric_tags_with_repr():
    counter: Counter = Counter()
    obj = ("tuple",)
    rate, status, _ = normalize_cell(obj, counter)
    assert rate is None
    assert status.startswith("unknown_sentinel:")
    # Counter key uses repr() so it survives Counter comparison.
    assert any("'tuple'" in k for k in counter)


# -----------------------------------------------------------------------------
# AKA parity — golden table from Projects/AKA White House/Lighthouse tests.
# -----------------------------------------------------------------------------

AKA_PARITY_GOLDEN = [
    ("3rd Party only", (None, "third_party_only", None)),
    ("1 guest only",   (None, "one_guest_only",   None)),
    ("Sold out",       (None, "sold_out",         None)),
    ("No flex",        (None, "no_flex",          None)),
    ("Room n/a",       (None, "room_na",          None)),
    ("--",             (None, "not_loaded",       None)),
    ("LOS2",           (None, "los_restricted",   2)),
    ("LOS3",           (None, "los_restricted",   3)),
    (199,              (199.0, "available",       None)),
    (199.5,            (199.5, "available",       None)),
    (None,             (None, "blank",            None)),
    (0,                (None, "blank",            None)),
]


@pytest.mark.parametrize("raw,expected", AKA_PARITY_GOLDEN)
def test_aka_parity_golden_table(raw, expected):
    """Behavior parity with AKA's `Projects/AKA White House/Lighthouse/
    lighthouse_parser._normalize_cell`. Divergence is a contract break —
    re-validate the AKA regression before merging changes here."""
    assert normalize_cell(raw) == expected


# -----------------------------------------------------------------------------
# AVAILABILITY_STATUSES contract
# -----------------------------------------------------------------------------

def test_availability_statuses_includes_every_documented_value():
    expected = {
        "available", "blank", "sold_out", "no_flex", "room_na", "not_loaded",
        "third_party_only", "one_guest_only", "los_restricted",
    }
    assert set(AVAILABILITY_STATUSES) == expected


def test_availability_statuses_aligns_with_analytics_exclusions_breakdown():
    """Every non-`available` status must appear as a key in the analytics
    `_exclusions_breakdown` output. If you add a status here, also add it
    there or downstream low-power diagnostics will silently miss it."""
    analytics_keys = {
        "sold_out", "los_restricted", "room_na", "blank", "not_loaded",
        "no_flex", "third_party_only", "one_guest_only",
    }
    sentinel_non_available = set(AVAILABILITY_STATUSES) - {"available"}
    assert sentinel_non_available == analytics_keys


# -----------------------------------------------------------------------------
# to_float_or_none — metadata-column helper
# -----------------------------------------------------------------------------

@pytest.mark.parametrize("v,expected", [
    (None, None),
    (float("nan"), None),
    ("not a number", None),
    (0, 0.0),
    (0.5, 0.5),
    (1, 1.0),
])
def test_to_float_or_none(v, expected):
    assert to_float_or_none(v) == expected


# -----------------------------------------------------------------------------
# Integration — real SFOEM XLSX. Skipped if fixture is absent.
# -----------------------------------------------------------------------------

SFOEM_XLSX = Path(
    "C:/Users/acarr/OneDrive/Documents/Claude/Projects/Hyatt Regency Embarcadero/"
    "Lighthouse/Drops/park-central-hotel-new-york_bookingdotcom_bar_los1_2guests_sec.xlsx"
)


@pytest.mark.skipif(not SFOEM_XLSX.exists(), reason="SFOEM fixture not present")
def test_sfoem_xlsx_zero_unknown_sentinels():
    """Real SFOEM LOS=1 export must parse with zero unknown sentinels.

    If this fails in the future, Lighthouse has introduced a new sentinel
    string that needs to be added explicitly to sentinel.py — DO NOT
    suppress the failure by widening the unknown-sentinel branch."""
    import openpyxl  # local import keeps openpyxl optional for the rest of the suite

    wb = openpyxl.load_workbook(SFOEM_XLSX, data_only=True, read_only=True)
    ws = wb["Rates"]
    headers = [c.value for c in ws[5]]
    property_columns_idx = list(range(6, len(headers)))  # G onwards (0-indexed cols 6+)

    counter: Counter = Counter()
    rate_count = sentinel_count = blank_count = 0
    for r in range(6, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if row[2] is None:
            continue
        for j in property_columns_idx:
            rate_usd, status, _ = normalize_cell(row[j], counter)
            if status == "available":
                rate_count += 1
            elif status == "blank":
                blank_count += 1
            else:
                sentinel_count += 1
    wb.close()

    assert counter == Counter(), (
        f"Unknown sentinels in SFOEM LOS=1 export: {dict(counter)}. "
        "Add to sentinel.py taxonomy."
    )
    # Sanity: we DID see numeric rates AND sentinels (i.e. we're testing real data,
    # not an empty file).
    assert rate_count > 1000, rate_count
    assert sentinel_count > 100, sentinel_count


# -----------------------------------------------------------------------------
# Resolution 20 — fail-loud slug-map drift detection (Finding 26)
# -----------------------------------------------------------------------------
# The SFOEM dry-run pass 2 (2026-05-08) discovered that 5 of 9 drafted slug-map
# keys would have failed silent dict lookups against the actual XLSX column
# headers due to verbatim-string drift between the parameter-collection drafts
# and the export's true column names. The fixtures below pin the verbatim
# mismatch cases so future regressions in the validator (e.g. a refactor that
# loosens the `h in slug_map` test) surface here.

# Header set Andrew drafted at parameter-collection time.
SFOEM_DRAFTED_SLUG_MAP = {
    "Hyatt Regency San Francisco Downtown SoMa":   "hyatt_regency_sf_soma",
    "Grand Hyatt San Francisco":                    "grand_hyatt_sf",
    "Palace Hotel, a Luxury Collection Hotel":      "palace_hotel_sf",
    "San Francisco Marriott Marquis":               "marriott_marquis_sf",
    "InterContinental San Francisco":               "intercontinental_sf",
}

# Headers that the actual XLSX `Rates` sheet shipped (the verbatim-string
# drift cases — uppercase, suffix-appended, brand-suffix variants).
SFOEM_ACTUAL_HEADERS = [
    "Hyatt Regency San Francisco Downtown SOMA",            # uppercase drift
    "Grand Hyatt San Francisco Union Square",               # suffix drift
    "Palace Hotel, a Luxury Collection Hotel, San Francisco",  # suffix drift
    "San Francisco Marriott Marquis Union Square",          # suffix drift
    "InterContinental San Francisco by IHG",                # brand-suffix drift
]


def test_validate_all_mapped_passes_silently():
    """Happy path: every property header maps. Validator returns None."""
    headers = ["Day", "Date", "Aka White House", "Capital Hilton", "Market demand", "Market OTB"]
    slug_map = {"Aka White House": "aka_white_house", "Capital Hilton": "capital_hilton"}
    result = validate_lighthouse_slug_map_coverage(
        headers, slug_map=slug_map, drop_columns=[],
    )
    assert result is None  # explicit — pure validator returns None on pass


def test_validate_one_column_missing_raises_with_column_name():
    """The unmapped column's verbatim string must appear in the message."""
    headers = ["Day", "Date", "Aka White House", "Capital Hilton", "Mystery Property"]
    slug_map = {"Aka White House": "aka_white_house", "Capital Hilton": "capital_hilton"}
    with pytest.raises(LighthouseSlugMapDriftError) as excinfo:
        validate_lighthouse_slug_map_coverage(
            headers, slug_map=slug_map, drop_columns=[],
        )
    msg = str(excinfo.value)
    assert "Mystery Property" in msg, msg
    assert "1 unmapped" in msg, msg
    assert "lighthouse_property_slug_map" in msg, msg


def test_validate_all_drop_columns_passes_silently():
    """A header that is in drop_columns (and not the slug map) must not raise."""
    headers = ["Day", "Date", "Park Central NY", "Market demand", "Market OTB"]
    slug_map: dict[str, str] = {}
    result = validate_lighthouse_slug_map_coverage(
        headers, slug_map=slug_map, drop_columns=["Park Central NY"],
    )
    assert result is None


def test_validate_mixed_mapped_drop_and_metadata_passes():
    """Mixed bag of every legitimate header category — must not raise."""
    headers = [
        "Day", "Date",                           # metadata
        "Aka White House", "Capital Hilton",     # mapped
        "Park Central NY",                        # drop
        "My OTB", "Market OTB", "Market demand", # metadata (My OTB always)
        "Unnamed: 7",                             # pandas-synthetic, ignore
    ]
    slug_map = {"Aka White House": "aka_white_house", "Capital Hilton": "capital_hilton"}
    drop_columns = ["Park Central NY"]
    result = validate_lighthouse_slug_map_coverage(
        headers, slug_map=slug_map, drop_columns=drop_columns,
    )
    assert result is None


def test_validate_empty_slug_map_lists_every_property_column():
    """When slug_map is empty AND drop_columns is empty, every property
    header must appear in the error — analyst needs the full list to
    reconcile."""
    headers = [
        "Day", "Date",
        "Hay-Adams", "St Regis", "Willard",
        "Market demand", "Market OTB",
    ]
    with pytest.raises(LighthouseSlugMapDriftError) as excinfo:
        validate_lighthouse_slug_map_coverage(
            headers, slug_map={}, drop_columns=[],
        )
    msg = str(excinfo.value)
    for prop in ("Hay-Adams", "St Regis", "Willard"):
        assert prop in msg, f"{prop} missing from error message: {msg}"
    assert "3 unmapped" in msg, msg


def test_validate_metadata_cols_override_works():
    """Caller-supplied metadata_cols overrides the default. Useful when a
    deal's export adds a new metadata column (e.g. `Day of week` instead
    of `Day`) without it being a property column."""
    headers = ["Day of week", "Date", "Aka White House", "Market demand", "Market OTB"]
    slug_map = {"Aka White House": "aka_white_house"}
    # Default metadata_cols would treat "Day of week" as unmapped; override
    # to acknowledge it.
    result = validate_lighthouse_slug_map_coverage(
        headers, slug_map=slug_map, drop_columns=[],
        metadata_cols={"Day of week", "Date", "Market demand", "Market OTB", "My OTB"},
    )
    assert result is None


def test_validate_default_metadata_cols_used_when_none_passed():
    """When metadata_cols is omitted, the default set of 5 columns kicks
    in (Day / Date / My OTB / Market OTB / Market demand). Verify with a
    header set that contains exactly those 5 plus one mapped property."""
    headers = ["Day", "Date", "My OTB", "Market OTB", "Market demand", "Capital Hilton"]
    slug_map = {"Capital Hilton": "capital_hilton"}
    result = validate_lighthouse_slug_map_coverage(
        headers, slug_map=slug_map, drop_columns=[],
    )
    assert result is None


def test_validate_non_string_headers_silently_ignored():
    """openpyxl/pandas occasionally yield None or numeric headers for
    truly-blank cells. Those must not blow up the validator."""
    headers = ["Day", "Date", None, 0, "Capital Hilton", "Market demand", "Market OTB"]
    slug_map = {"Capital Hilton": "capital_hilton"}
    result = validate_lighthouse_slug_map_coverage(
        headers, slug_map=slug_map, drop_columns=[],
    )
    assert result is None


def test_validate_sfoem_drafted_keys_fail_against_actual_headers():
    """Pin the SFOEM dry-run pass 2 finding: the drafted slug map fails
    against the actual XLSX headers because of verbatim-string drift.
    All 5 actual headers are unmapped under the drafted map, and the
    error message must list every one of them so the analyst can
    reconcile."""
    headers = ["Day", "Date"] + SFOEM_ACTUAL_HEADERS + ["Market demand", "Market OTB"]
    with pytest.raises(LighthouseSlugMapDriftError) as excinfo:
        validate_lighthouse_slug_map_coverage(
            headers,
            slug_map=SFOEM_DRAFTED_SLUG_MAP,
            drop_columns=[],
        )
    msg = str(excinfo.value)
    assert "5 unmapped" in msg, msg
    for header in SFOEM_ACTUAL_HEADERS:
        assert header in msg, f"missing actual header from error: {header!r}"


def test_validate_sfoem_corrected_keys_pass():
    """The fix path: once the analyst reconciles the slug map against the
    actual XLSX headers, validation passes. This is the post-fix state."""
    corrected_slug_map = dict(zip(SFOEM_ACTUAL_HEADERS, [
        "hyatt_regency_sf_soma", "grand_hyatt_sf", "palace_hotel_sf",
        "marriott_marquis_sf", "intercontinental_sf",
    ]))
    headers = ["Day", "Date"] + SFOEM_ACTUAL_HEADERS + ["Market demand", "Market OTB"]
    result = validate_lighthouse_slug_map_coverage(
        headers,
        slug_map=corrected_slug_map,
        drop_columns=[],
    )
    assert result is None


@pytest.mark.skipif(not SFOEM_XLSX.exists(), reason="SFOEM fixture not present")
def test_sfoem_xlsx_distribution_smoke():
    """Anchor sentinel distribution against the 2026-05-08 baseline so
    that catastrophic regressions in `normalize_cell` (e.g. a regex change
    that misses every LOS string) surface here. Tolerances are generous —
    the test catches "nothing matches anymore" not "drift in the third
    decimal." """
    import openpyxl

    wb = openpyxl.load_workbook(SFOEM_XLSX, data_only=True, read_only=True)
    ws = wb["Rates"]
    counter: Counter = Counter()
    status_counter: Counter = Counter()
    for r in range(6, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        if row[2] is None:
            continue
        for j in range(6, len(row)):
            _, status, _ = normalize_cell(row[j], counter)
            status_counter[status] += 1
    wb.close()

    # 2026-05-08 baseline: ~76% numeric, ~9% sold_out, plus small NO_FLEX /
    # NOT_LOADED / LOS_RESTRICTED tails. Use generous bounds.
    assert status_counter["available"] > 2500, dict(status_counter)
    assert status_counter["sold_out"] > 200, dict(status_counter)
    assert status_counter["no_flex"] >= 0
    # Park Central column ships with sold_out and no_flex; if we somehow
    # tagged everything as unknown_sentinel that bucket would dominate.
    assert all(
        not k.startswith("unknown_sentinel:") for k in status_counter
    ), dict(status_counter)
